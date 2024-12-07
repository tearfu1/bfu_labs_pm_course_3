import re

import openpyxl


class Employee:
	def __init__(self, salary):
		self.tab_number = ""
		self.last_name = ""
		self.dept = ""
		self.salary_before = salary
		self.salary_after = salary

	def fill(self, tab_number, last_name, dept, salary_before, salary_after):
		self.tab_number = tab_number
		self.last_name = last_name
		self.dept = dept
		self.salary_before = salary_before
		self.salary_after = salary_after


def parse_currency(value):
	cleaned_value = re.sub(r"[^\d,-]", "", value)
	cleaned_value = cleaned_value.replace(",", ".")
	return float(cleaned_value)


def is_total_cell(value):
	return "Итог" in value


def main():
	path = "Report.xlsx"
	wb = openpyxl.load_workbook(path)
	ws = wb.active
	ws.title = "Report"

	max_row = ws.max_row
	max_col = ws.max_column

	avg_salaries = {}
	employee_cnt = 0
	max_salary_emplyee = Employee(0)
	min_salary_emplyee = Employee(float("inf"))

	for row in ws.iter_rows(min_row=2, max_row=max_row - 1, min_col=1, max_col=max_col):
		row_values = [cell.value for cell in row]

		tab_num = row_values[0]
		last_name = row_values[1]
		dept = row_values[2]
		salary_before = parse_currency(row_values[5])
		salary_after = parse_currency(row_values[8])

		if is_total_cell(dept):
			dept_salaries = [salary_before, salary_after]
			clear_dept = dept[:-5]
			avg_salaries[clear_dept] = list(map(lambda x: round(x / employee_cnt, 2), dept_salaries))
			employee_cnt = 0
			continue

		if salary_before > max_salary_emplyee.salary_before:
			max_salary_emplyee.fill(tab_num, last_name, dept, salary_before, salary_after)
		if salary_before < min_salary_emplyee.salary_before:
			min_salary_emplyee.fill(tab_num, last_name, dept, salary_before, salary_after)

		employee_cnt += 1

	ws.append([""])
	headers = ["Таб. номер", "Фамилия", "Отдел", "Зарплата до вычета", "Зарплата после вычета"]
	ws.append(headers)
	ws.append(list(max_salary_emplyee.__dict__.values()))
	ws.append(list(min_salary_emplyee.__dict__.values()))

	ws.append([""])
	headers = ["Отдел", "Ср. Зарплата до вычета", "Ср. Зарплата после вычета"]
	ws.append(headers)
	for dept, salaries in avg_salaries.items():
		ws.append([dept, salaries[0], salaries[1]])

	output_path = "Report.xlsx"
	wb.save(output_path)


if __name__ == "__main__":
	main()
	print("task2 done")
