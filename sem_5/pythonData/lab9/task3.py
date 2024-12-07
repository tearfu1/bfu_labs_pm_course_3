import openpyxl
from openpyxl.chart import PieChart, Reference

from task1 import format_currency
from task2 import is_total_cell, parse_currency


def main():
	path = "Report.xlsx"
	wb = openpyxl.load_workbook(path)
	ws = wb.active

	max_row = 11
	max_col = ws.max_column

	dept_totals = {}

	for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col):
		row_values = [cell.value for cell in row]
		dept = row_values[2]
		if is_total_cell(dept):
			salary_before = parse_currency(row_values[5])
			clear_dept = dept[:-5]
			dept_totals[clear_dept] = salary_before

	total_count_row = max_row + 1
	total_count_col_before = chr(ord('A') + 5)
	overall_totals = parse_currency(ws[f"{total_count_col_before}{total_count_row}"].value)

	dept_totals_parts = {}
	for key, value in dept_totals.items():
		dept_totals_parts[key] = round(value / overall_totals, 2)

	ws.append([""])
	headers = ["Отдел", "Доля от общей суммы"]
	ws.append(headers)

	total_row = ws.max_row
	for dept, value in dept_totals_parts.items():
		ws.append([dept, value])
	pie = PieChart()
	labels = Reference(ws,
					   min_col=1,
					   min_row=total_row + 1,
					   max_row=total_row + len(dept_totals_parts),
					   max_col=1)

	data = Reference(ws,
					 min_col=2,
					 min_row=total_row,
					 max_row=total_row + len(dept_totals_parts),
					 max_col=2)

	pie.add_data(data, titles_from_data=True)
	pie.set_categories(labels)
	pie.title = "Зарплата по отделам"

	ws.add_chart(pie, "K2")

	output_path = "Report.xlsx"
	wb.save(output_path)


if __name__ == "__main__":
	main()
	print("task3 done")
