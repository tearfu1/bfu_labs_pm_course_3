from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def calculate_fields(record):
	ndfl = (record["НДФЛ"] * record["Сумма зарплаты"]) / 100
	to_pay = record["Сумма зарплаты"] - ndfl
	return round(ndfl, 2), round(to_pay, 2)


def format_currency(value):
	value = round(value, 2)
	integer_part, decimal_part = str(value).split(".")

	integer_part = "{:,}".format(int(integer_part)).replace(",", " ")

	return f"{integer_part},{decimal_part}р."


def format_percent(value):
	return f"{value}%"


def format_total_title(value):
	return f"{value} Итог"


def ws_append_totals(ws, totals, dept="Общий"):
	ws.append(["",
			   "",
			   format_total_title(dept),
			   format_currency(totals["Сумма по окладу"]),
			   format_currency(totals["Сумма по надбавкам"]),
			   format_currency(totals["Сумма зарплаты"]),
			   "",
			   format_currency(totals["Сумма НДФЛ"]),
			   format_currency(totals["Сумма к выдаче"])])


def main():
	wb = Workbook()
	ws = wb.active
	ws.title = "Report"

	headers = ["Таб. номер", "Фамилия", "Отдел", "Сумма по окладу, руб.", "Сумма по надбавкам, руб.",
			   "Сумма зарплаты, руб.", "НДФЛ, %", "Сумма НДФЛ, руб.", "Сумма к выдаче, руб."]
	ws.append(headers)

	all_cols_letters = {'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I'}

	for col_letter in all_cols_letters:
		ws.column_dimensions[col_letter].width = 22

	ws.row_dimensions[1].height = 100

	for col_num, header in enumerate(headers, start=1):
		col_letter = get_column_letter(col_num)
		ws[f"{col_letter}1"].alignment = Alignment(horizontal="center", vertical="center", textRotation=90,
												   wrapText=True)
	data = [
		{"Таб. номер": "0002", "Фамилия": "Петров П.П.", "Отдел": "Бухгалтерия",
		 "Сумма по окладу": 3913.04, "Сумма по надбавкам": 2608.70, "Сумма зарплаты": 6521.74, "НДФЛ": 13},

		{"Таб. номер": "0005", "Фамилия": "Васин В.В.", "Отдел": "Бухгалтерия",
		 "Сумма по окладу": 5934.78, "Сумма по надбавкам": 913.04, "Сумма зарплаты": 6847.83, "НДФЛ": 13},

		{"Таб. номер": "0001", "Фамилия": "Иванов И.И.", "Отдел": "Отдел кадров",
		 "Сумма по окладу": 6000.00, "Сумма по надбавкам": 4000.00, "Сумма зарплаты": 10000.00, "НДФЛ": 13},

		{"Таб. номер": "0003", "Фамилия": "Сидоров С.С.", "Отдел": "Отдел кадров",
		 "Сумма по окладу": 5000.00, "Сумма по надбавкам": 4500.00, "Сумма зарплаты": 9500.00, "НДФЛ": 13},

		{"Таб. номер": "0006", "Фамилия": "Львов Л.Л.", "Отдел": "Отдел кадров",
		 "Сумма по окладу": 4074.07, "Сумма по надбавкам": 2444.44, "Сумма зарплаты": 6518.52, "НДФЛ": 13},

		{"Таб. номер": "0007", "Фамилия": "Волков В.В.", "Отдел": "Отдел кадров",
		 "Сумма по окладу": 1434.78, "Сумма по надбавкам": 1434.78, "Сумма зарплаты": 2869.57, "НДФЛ": 13},

		{"Таб. номер": "0004", "Фамилия": "Мишин М.М.", "Отдел": "Столовая",
		 "Сумма по окладу": 5500.00, "Сумма по надбавкам": 3500.00, "Сумма зарплаты": 9000.00, "НДФЛ": 13}
	]

	prev_dept = data[0]["Отдел"]
	totals = {"Сумма по окладу": 0, "Сумма по надбавкам": 0, "Сумма зарплаты": 0, "Сумма НДФЛ": 0,
			  "Сумма к выдаче": 0}
	overall_totals = {"Сумма по окладу": 0, "Сумма по надбавкам": 0, "Сумма зарплаты": 0, "Сумма НДФЛ": 0,
					  "Сумма к выдаче": 0}

	totals_keys = ["Сумма по окладу", "Сумма по надбавкам", "Сумма зарплаты"]
	row_index = 2

	for record in data:
		if prev_dept != record["Отдел"]:
			ws_append_totals(ws, totals, prev_dept)
			prev_dept = record["Отдел"]
			totals = {"Сумма по окладу": 0,
					  "Сумма по надбавкам": 0,
					  "Сумма зарплаты": 0,
					  "Сумма НДФЛ": 0,
					  "Сумма к выдаче": 0,
					  }
			ws[f"C{row_index}"].font = Font(bold=True)
			row_index += 1

		ndfl, to_pay = calculate_fields(record)
		ws.append([record["Таб. номер"],
				   record["Фамилия"],
				   record["Отдел"],
				   format_currency(record["Сумма по окладу"]),
				   format_currency(record["Сумма по надбавкам"]),
				   format_currency(record["Сумма зарплаты"]),
				   format_percent(record["НДФЛ"]),
				   format_currency(ndfl),
				   format_currency(to_pay)])

		for key in totals_keys:
			totals[key] += record[key]
			overall_totals[key] += record[key]

		totals["Сумма НДФЛ"] += ndfl
		totals["Сумма к выдаче"] += to_pay

		overall_totals["Сумма НДФЛ"] += ndfl
		overall_totals["Сумма к выдаче"] += to_pay

		row_index += 1

	ws_append_totals(ws, totals, prev_dept)

	ws[f"C{row_index}"].font = Font(bold=True)
	row_index += 1

	ws_append_totals(ws, overall_totals)

	ws[f"C{row_index}"].font = Font(bold=True)

	output_path = "Report.xlsx"
	wb.save(output_path)


if __name__ == "__main__":
	main()
	print("task1 done")
